from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from datetime import timedelta, date
import io

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

from sales.models import Sale, SaleItem
from inventory.models import Product, Stock


def _get_sales_qs(date_from, date_to, joint_pk=None):
    qs = Sale.objects.filter(
        is_held=False,
        sale_date__date__gte=date_from,
        sale_date__date__lte=date_to,
    ).select_related('joint', 'sold_by').prefetch_related('items__product')
    if joint_pk:
        qs = qs.filter(joint_id=joint_pk)
    return qs


@login_required
def weekly_excel_report(request):
    """Generate an Excel report for the current week (Mon–Sun)."""
    if not OPENPYXL_AVAILABLE:
        return HttpResponse("openpyxl not installed. Run: pip install openpyxl", status=500)

    today = date.today()
    monday = today - timedelta(days=today.weekday())
    sunday = monday + timedelta(days=6)

    joint_pk = request.GET.get('joint') or None
    sales = _get_sales_qs(monday, today, joint_pk)

    # ── Build Workbook ─────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Weekly Sales"

    # Colour palette
    DARK  = "0A0A0F"
    ACCENT = "6C47FF"
    GREEN  = "00D68F"
    LIGHT  = "F8F7FC"
    WHITE  = "FFFFFF"

    # Title row
    ws.merge_cells("A1:H1")
    ws["A1"] = f"GenX POS — Weekly Sales Report"
    ws["A1"].font = Font(name="Calibri", bold=True, size=16, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=DARK)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Week: {monday.strftime('%d %b')} – {sunday.strftime('%d %b %Y')}"
    ws["A2"].font = Font(name="Calibri", size=11, color="AAAAAA")
    ws["A2"].fill = PatternFill("solid", fgColor=DARK)
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 22

    ws.append([])  # blank

    # ── SECTION: SUMMARY STATS ─────────────────────────────────────
    ws.merge_cells("A4:H4")
    ws["A4"] = "Summary"
    ws["A4"].font = Font(bold=True, color=WHITE, size=12)
    ws["A4"].fill = PatternFill("solid", fgColor=ACCENT)
    ws["A4"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[4].height = 24

    total_rev = sum(s.total_amount for s in sales)
    total_count = sales.count()
    avg_val = (total_rev / total_count) if total_count else 0

    from django.db.models import Sum, Count
    payment_breakdown = {}
    for s in sales:
        m = s.get_payment_method_display()
        payment_breakdown[m] = payment_breakdown.get(m, 0) + 1

    stats = [
        ["Total Revenue", f"${total_rev:,.2f}"],
        ["Total Sales", str(total_count)],
        ["Average Sale Value", f"${avg_val:,.2f}"],
    ]
    for method, count in payment_breakdown.items():
        stats.append([f"  {method} Payments", str(count)])

    for row_data in stats:
        r = ws.max_row + 1
        ws.cell(r, 1, row_data[0]).font = Font(name="Calibri", size=11, color="444444")
        ws.cell(r, 2, row_data[1]).font = Font(name="Calibri", bold=True, size=11)
        ws.row_dimensions[r].height = 18

    ws.append([])

    # ── SECTION: SALES TABLE ───────────────────────────────────────
    header_row = ws.max_row + 1
    headers = ["Receipt #", "Date", "Joint", "Customer", "Staff", "Payment", "Items", "Total"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(header_row, col, h)
        cell.font = Font(bold=True, color=WHITE, size=10)
        cell.fill = PatternFill("solid", fgColor="1A1A2E")
        cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[header_row].height = 22

    thin = Side(style="thin", color="E0DEEC")
    border = Border(bottom=thin)

    for i, sale in enumerate(sales):
        r = ws.max_row + 1
        row_fill = PatternFill("solid", fgColor="FAF9FD" if i % 2 == 0 else WHITE)
        data = [
            sale.receipt_number,
            sale.sale_date.strftime("%d/%m/%Y %H:%M"),
            sale.joint.display_name,
            sale.customer_name or "—",
            sale.sold_by.get_full_name() or sale.sold_by.username,
            sale.get_payment_method_display(),
            sale.items.count(),
            sale.total_amount,
        ]
        for col, val in enumerate(data, 1):
            cell = ws.cell(r, col, val)
            cell.fill = row_fill
            cell.border = border
            cell.font = Font(name="Calibri", size=10)
            if col == 8:  # Total
                cell.number_format = '"$"#,##0.00'
                cell.font = Font(name="Calibri", bold=True, size=10, color="065F46")
            if col == 1:  # Receipt
                cell.font = Font(name="Courier New", size=10)

    # Total row
    r = ws.max_row + 1
    ws.cell(r, 7, "TOTAL").font = Font(bold=True, size=11)
    ws.cell(r, 8, total_rev).number_format = '"$"#,##0.00'
    ws.cell(r, 8).font = Font(bold=True, size=12, color="6C47FF")
    ws.cell(r, 8).fill = PatternFill("solid", fgColor="F0EEFF")

    # Column widths
    col_widths = [16, 18, 16, 20, 20, 14, 8, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── SECTION: TOP PRODUCTS ──────────────────────────────────────
    ws.append([])
    ws.append([])
    r = ws.max_row
    ws.merge_cells(f"A{r}:H{r}")
    ws.cell(r, 1, "Top Products This Week").font = Font(bold=True, color=WHITE, size=12)
    ws.cell(r, 1).fill = PatternFill("solid", fgColor=ACCENT)
    ws.cell(r, 1).alignment = Alignment(indent=1)
    ws.row_dimensions[r].height = 24

    from django.db.models import Sum as DSum
    top_products = (
        SaleItem.objects
        .filter(sale__in=sales, is_free_gift=False)
        .values('product__name', 'product__joint__display_name')
        .annotate(qty=DSum('quantity'), revenue=DSum('line_total'))
        .order_by('-qty')[:10]
    )

    r = ws.max_row + 1
    for h, col in [("Product", 1), ("Joint", 2), ("Qty Sold", 3), ("Revenue", 4)]:
        ws.cell(r, col, h).font = Font(bold=True, color=WHITE, size=10)
        ws.cell(r, col).fill = PatternFill("solid", fgColor="1A1A2E")
    ws.row_dimensions[r].height = 20

    for p in top_products:
        r = ws.max_row + 1
        ws.cell(r, 1, p['product__name']).font = Font(name="Calibri", bold=True, size=10)
        ws.cell(r, 2, p['product__joint__display_name']).font = Font(name="Calibri", size=10, color="666666")
        ws.cell(r, 3, p['qty']).font = Font(name="Calibri", bold=True, size=10, color="6C47FF")
        ws.cell(r, 4, p['revenue']).number_format = '"$"#,##0.00'
        ws.cell(r, 4).font = Font(name="Calibri", bold=True, size=10, color="065F46")

    # Output
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"genx-weekly-report-{monday.strftime('%Y-%m-%d')}.xlsx"
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def monthly_excel_report(request):
    """Generate an Excel report for the current or specified month."""
    if not OPENPYXL_AVAILABLE:
        return HttpResponse("openpyxl not installed. Run: pip install openpyxl", status=500)

    today = date.today()
    year  = int(request.GET.get('year',  today.year))
    month = int(request.GET.get('month', today.month))
    import calendar
    _, last_day = calendar.monthrange(year, month)
    date_from = date(year, month, 1)
    date_to   = date(year, month, last_day)
    month_name = date_from.strftime("%B %Y")

    joint_pk = request.GET.get('joint') or None
    sales = _get_sales_qs(date_from, date_to, joint_pk)

    # Re-use same logic as weekly but monthly title
    # (For brevity, call the same builder with different dates)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Monthly Sales"
    DARK = "0A0A0F"; ACCENT = "6C47FF"; WHITE = "FFFFFF"

    ws.merge_cells("A1:H1")
    ws["A1"] = f"GenX POS — Monthly Report: {month_name}"
    ws["A1"].font = Font(name="Calibri", bold=True, size=16, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=DARK)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    total_rev = sum(s.total_amount for s in sales)
    total_count = sales.count()
    avg_val = (total_rev / total_count) if total_count else 0

    ws.append([])
    ws.cell(ws.max_row, 1, "Metric").font = Font(bold=True, color=WHITE)
    ws.cell(ws.max_row, 1).fill = PatternFill("solid", fgColor=ACCENT)
    ws.cell(ws.max_row, 2, "Value").font = Font(bold=True, color=WHITE)
    ws.cell(ws.max_row, 2).fill = PatternFill("solid", fgColor=ACCENT)

    for label, value in [
        ("Total Revenue", f"${total_rev:,.2f}"),
        ("Total Transactions", str(total_count)),
        ("Average Transaction Value", f"${avg_val:,.2f}"),
    ]:
        r = ws.max_row + 1
        ws.cell(r, 1, label).font = Font(size=11)
        ws.cell(r, 2, value).font = Font(bold=True, size=11)

    ws.append([])

    header_row = ws.max_row + 1
    for col, h in enumerate(["Receipt #", "Date", "Joint", "Customer", "Staff", "Payment", "Items", "Total"], 1):
        c = ws.cell(header_row, col, h)
        c.font = Font(bold=True, color=WHITE, size=10)
        c.fill = PatternFill("solid", fgColor="1A1A2E")

    for i, sale in enumerate(sales):
        r = ws.max_row + 1
        fill = PatternFill("solid", fgColor="FAF9FD" if i % 2 == 0 else WHITE)
        for col, val in enumerate([
            sale.receipt_number,
            sale.sale_date.strftime("%d/%m/%Y %H:%M"),
            sale.joint.display_name,
            sale.customer_name or "—",
            sale.sold_by.get_full_name() or sale.sold_by.username,
            sale.get_payment_method_display(),
            sale.items.count(),
            sale.total_amount,
        ], 1):
            c = ws.cell(r, col, val)
            c.fill = fill
            c.font = Font(name="Calibri", size=10)
            if col == 8:
                c.number_format = '"$"#,##0.00'
                c.font = Font(name="Calibri", bold=True, size=10, color="065F46")

    for i, w in enumerate([16,18,14,18,18,12,8,14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"genx-monthly-report-{date_from.strftime('%Y-%m')}.xlsx"
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def monthly_pdf_report(request):
    """Generate a PDF monthly report using ReportLab."""
    if not REPORTLAB_AVAILABLE:
        return HttpResponse("reportlab not installed. Run: pip install reportlab", status=500)

    today = date.today()
    year  = int(request.GET.get('year',  today.year))
    month = int(request.GET.get('month', today.month))
    import calendar
    _, last_day = calendar.monthrange(year, month)
    date_from = date(year, month, 1)
    date_to   = date(year, month, last_day)
    month_name = date_from.strftime("%B %Y")

    joint_pk = request.GET.get('joint') or None
    sales = _get_sales_qs(date_from, date_to, joint_pk)

    total_rev   = sum(s.total_amount for s in sales)
    total_count = sales.count()
    avg_val     = (total_rev / total_count) if total_count else 0

    # ── Build PDF ──────────────────────────────────────────────────
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=2*cm, rightMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    ink_color  = colors.HexColor("#0A0A0F")
    accent_col = colors.HexColor("#6C47FF")
    green_col  = colors.HexColor("#00D68F")
    muted_col  = colors.HexColor("#888888")

    title_style = ParagraphStyle('Title', parent=styles['Normal'],
        fontName='Helvetica-Bold', fontSize=22, textColor=ink_color,
        spaceAfter=4, alignment=TA_CENTER)
    sub_style = ParagraphStyle('Sub', parent=styles['Normal'],
        fontName='Helvetica', fontSize=11, textColor=muted_col,
        spaceAfter=20, alignment=TA_CENTER)
    section_style = ParagraphStyle('Section', parent=styles['Normal'],
        fontName='Helvetica-Bold', fontSize=13, textColor=accent_col,
        spaceAfter=8, spaceBefore=16)
    label_style = ParagraphStyle('Label', parent=styles['Normal'],
        fontName='Helvetica', fontSize=10, textColor=muted_col)

    elements = []

    # Header
    elements.append(Paragraph("GenX POS", title_style))
    elements.append(Paragraph(f"Monthly Sales Report — {month_name}", sub_style))
    elements.append(Spacer(1, 0.3*cm))

    # Summary table
    elements.append(Paragraph("Summary", section_style))
    summary_data = [
        ["Metric", "Value"],
        ["Total Revenue", f"${total_rev:,.2f}"],
        ["Total Transactions", str(total_count)],
        ["Average Transaction Value", f"${avg_val:,.2f}"],
    ]
    summary_table = Table(summary_data, colWidths=[9*cm, 8*cm])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), accent_col),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,0), 10),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor("#FAF9FD"), colors.white]),
        ('FONTNAME',   (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE',   (0,1), (-1,-1), 10),
        ('FONTNAME',   (1,1), (1,-1), 'Helvetica-Bold'),
        ('ALIGN',      (1,0), (1,-1), 'RIGHT'),
        ('ROWHEIGHT',  (0,0), (-1,-1), 22),
        ('LEFTPADDING',  (0,0), (-1,-1), 10),
        ('RIGHTPADDING', (0,0), (-1,-1), 10),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E0DEEC")),
        ('ROUNDEDCORNERS', [4]),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.5*cm))

    # Sales table
    elements.append(Paragraph("Transaction Detail", section_style))
    sale_headers = ["Receipt", "Date", "Joint", "Staff", "Payment", "Total"]
    sale_rows = [sale_headers]
    for sale in sales[:100]:  # Cap at 100 rows for PDF length
        sale_rows.append([
            sale.receipt_number,
            sale.sale_date.strftime("%d/%m/%Y"),
            sale.joint.display_name,
            sale.sold_by.get_full_name() or sale.sold_by.username,
            sale.get_payment_method_display(),
            f"${sale.total_amount:,.2f}",
        ])
    if sales.count() > 100:
        sale_rows.append(["... (truncated)", "", "", "", "", ""])

    sale_table = Table(sale_rows, colWidths=[3.2*cm, 2.5*cm, 3*cm, 3.5*cm, 2.5*cm, 2.5*cm])
    sale_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), ink_color),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,-1), 9),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor("#FAF9FD"), colors.white]),
        ('ALIGN',      (5,0), (5,-1), 'RIGHT'),
        ('FONTNAME',   (5,1), (5,-1), 'Helvetica-Bold'),
        ('TEXTCOLOR',  (5,1), (5,-1), green_col),
        ('GRID', (0,0), (-1,-1), 0.3, colors.HexColor("#E0DEEC")),
        ('LEFTPADDING',  (0,0), (-1,-1), 6),
        ('RIGHTPADDING', (0,0), (-1,-1), 6),
        ('ROWHEIGHT', (0,0), (-1,-1), 18),
    ]))
    elements.append(sale_table)

    # Footer
    elements.append(Spacer(1, 1*cm))
    elements.append(Paragraph(
        f"Generated {today.strftime('%d %B %Y')} — GenX POS · Harare, Zimbabwe",
        ParagraphStyle('Footer', parent=styles['Normal'],
            fontName='Helvetica', fontSize=8, textColor=muted_col, alignment=TA_CENTER)
    ))

    doc.build(elements)
    buffer.seek(0)

    filename = f"genx-monthly-report-{date_from.strftime('%Y-%m')}.pdf"
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response